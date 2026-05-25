#!/usr/bin/env python3
"""Build or update a portfolio roll-up from extracted fund data.

Usage:  python build_rollup.py <spec.json>

The roll-up is a living workbook with three tabs:
  - Portfolio Roll-Up   : current-state dashboard, one row per fund (latest figures)
  - History            : a frozen quarterly log -- per-fund rows + a portfolio
                          total row for every quarter the skill has been run
  - Reconciliation Key : audit trail tracing every current figure to its source

Each run REFRESHES the dashboard (updates a fund's row if a new statement was
provided, appends genuinely new funds, carries the rest forward unchanged) and
APPENDS this quarter's snapshot to History. Re-running for the same quarter
replaces that quarter's History block, so it is safe to run twice.

The spec JSON is assembled by Claude after reading the statements. See
references/reading-statements.md for the spec format and a worked example.
"""
import json
import math
import os
import re
import shutil
import subprocess
import sys
import tempfile
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties

NAVY = "FF1F3A5F"
SUBHEAD = "FF4A5F7A"
SHADE = "FFEEF1F5"
BLUE = "FF0000FF"      # figures keyed straight from a statement
GREEN = "FF008000"     # links to another sheet
CUR = '$#,##0'
MULT = '0.00"x"'
PCT = '0.0%'

BLUE_FIELDS = ["Vintage", "Commitment", "Paid-In Capital",
               "Cumulative Distributions", "Current NAV", "Net IRR"]

RECON_FIELDS = [
    ("Commitment", "Commitment", CUR),
    ("Paid-In Capital", "Paid-In Capital", CUR),
    ("Unfunded Commitment", "Unfunded Commitment", CUR),
    ("Cumulative Distributions", "Cumulative Distributions", CUR),
    ("Current NAV", "Current NAV", CUR),
    ("Total Value", "Total Value", CUR),
    ("DPI", "DPI", MULT),
    ("TVPI", "TVPI", MULT),
    ("Net IRR", "Net IRR", PCT),
]

HISTORY_COLS = ["Quarter", "Fund", "Strategy", "Commitment", "Paid-In Capital",
                "Unfunded Commitment", "Cumulative Distributions", "Current NAV",
                "Total Value", "DPI", "TVPI", "Net IRR"]
HISTORY_FMT = {"Commitment": CUR, "Paid-In Capital": CUR, "Unfunded Commitment": CUR,
               "Cumulative Distributions": CUR, "Current NAV": CUR, "Total Value": CUR,
               "DPI": MULT, "TVPI": MULT, "Net IRR": PCT}

COLLETTER = {}


def quarter_key(label):
    """Sort key so 'Q1 2026' < 'Q2 2026' < 'Q1 2027'."""
    m = re.search(r'Q\s*([1-4])\D+(\d{4})', str(label))
    if m:
        return (int(m.group(2)), int(m.group(1)))
    m = re.search(r'(\d{4})', str(label))
    return (int(m.group(1)), 0) if m else (9999, 9)


def find_rollup_sheet(wb):
    """Find the dashboard sheet (not History / Reconciliation), its header row,
    and a header-text -> column map."""
    candidates = [ws for ws in wb.worksheets
                  if "histor" not in ws.title.lower()
                  and "recon" not in ws.title.lower()]
    candidates.sort(key=lambda ws: 0 if re.search(r'roll.?up', ws.title, re.I) else 1)
    for ws in candidates:
        for row in ws.iter_rows():
            vals = [str(c.value).strip() if c.value is not None else "" for c in row]
            if "Fund" in vals and "Commitment" in vals and "Net IRR" in vals \
                    and "Strategy" in vals:
                col = {str(c.value).strip(): c.column for c in row if c.value is not None}
                return ws, row[0].row, col
    raise SystemExit("ERROR: could not find the Portfolio Roll-Up header row.")


def find_total_row(ws, hdr, col_a):
    for r in range(hdr + 1, ws.max_row + 2):
        v = ws.cell(row=r, column=col_a).value
        if v is not None and str(v).strip().upper().startswith("TOTAL"):
            return r
    raise SystemExit("ERROR: could not find the TOTAL row in the roll-up sheet.")


def recalc(path):
    """Round-trip through LibreOffice with recalc-on-load forced so cached
    formula values are correct. Best-effort -- the formulas are valid regardless."""
    try:
        prof = tempfile.mkdtemp(prefix="lo_prof_")
        ud = os.path.join(prof, "user")
        os.makedirs(ud, exist_ok=True)
        open(os.path.join(ud, "registrymodifications.xcu"), "w").write(
            '<?xml version="1.0" encoding="UTF-8"?>\n<oor:items '
            'xmlns:oor="http://openoffice.org/2001/registry" '
            'xmlns:xs="http://www.w3.org/2001/XMLSchema">\n'
            '<item oor:path="/org.openoffice.Office.Calc/Formula/Load">'
            '<prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>0</value></prop></item>\n'
            '</oor:items>\n')
        out = tempfile.mkdtemp(prefix="lo_out_")
        subprocess.run(["soffice", "-env:UserInstallation=file://" + prof, "--headless",
                        "--calc", "--convert-to", "xlsx", "--outdir", out, path],
                       check=True, capture_output=True, timeout=180)
        p = os.path.join(out, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
        if os.path.exists(p):
            shutil.move(p, path)
            return True
    except Exception as e:
        print(f"  (recalc skipped -- {e}; formulas are valid, Excel recalcs on open)")
    return False


def read_existing_funds(ws, hdr, COL, total_row):
    """Read the funds already on the dashboard, in row order."""
    out = []
    for r in range(hdr + 1, total_row):
        name = ws.cell(row=r, column=COL["Fund"]).value
        if name is None or str(name).strip() == "":
            continue
        def g(field):
            return ws.cell(row=r, column=COL[field]).value if field in COL else None
        out.append({
            "name": str(name).strip(),
            "strategy": g("Strategy") or "",
            "vintage": g("Vintage"),
            "commitment": g("Commitment") or 0,
            "paid_in": g("Paid-In Capital") or 0,
            "distributions": g("Cumulative Distributions") or 0,
            "nav": g("Current NAV") or 0,
            "net_irr": g("Net IRR"),
            "_fresh": False,
        })
    return out


def read_existing_history(wb):
    """Return {quarter: [fund-dict,...]} of frozen prior snapshots (fund rows only)."""
    if "History" not in wb.sheetnames:
        return {}
    hs = wb["History"]
    hdr = None
    for row in hs.iter_rows():
        vals = [str(c.value).strip() if c.value is not None else "" for c in row]
        if "Quarter" in vals and "Fund" in vals:
            hdr = row[0].row
            # Build the column map from the contiguous main-table header block
            # only. The chart-source helper table further right reuses some
            # header names ("Quarter", "Current NAV") -- stopping at the first
            # gap keeps those out of the map.
            cmap = {}
            for c in row:
                if c.value is None:
                    if cmap:
                        break
                    continue
                cmap.setdefault(str(c.value).strip(), c.column)
            break
    if hdr is None:
        return {}
    blocks = {}
    for r in range(hdr + 1, hs.max_row + 1):
        q = hs.cell(row=r, column=cmap["Quarter"]).value
        fund = hs.cell(row=r, column=cmap["Fund"]).value
        if not q or not fund or str(fund).strip().upper().startswith("TOTAL"):
            continue
        def g(field):
            return hs.cell(row=r, column=cmap[field]).value if field in cmap else None
        blocks.setdefault(str(q).strip(), []).append({
            "name": str(fund).strip(),
            "strategy": g("Strategy") or "",
            "commitment": g("Commitment") or 0,
            "paid_in": g("Paid-In Capital") or 0,
            "distributions": g("Cumulative Distributions") or 0,
            "nav": g("Current NAV") or 0,
            "net_irr": g("Net IRR"),
        })
    return blocks


def write_front_page(ws, hdr, COL, total_row, funds, spec):
    """Rewrite the dashboard fund block from the merged fund list."""
    n = len(funds)
    capacity = total_row - hdr - 1
    if n > capacity:
        ws.insert_rows(total_row, n - capacity)
        total_row += (n - capacity)
    first, last = hdr + 1, hdr + n

    fmt, fnt = {}, {}
    for name, c in COL.items():
        cell = ws.cell(row=first, column=c)
        fmt[name] = cell.number_format
        fnt[name] = copy(cell.font)

    def setcell(r, name, value, formula=False):
        c = COL[name]
        cell = ws.cell(row=r, column=c, value=value)
        cell.number_format = fmt[name]
        base = fnt[name]
        is_blue = (not formula) and name in BLUE_FIELDS
        cell.font = Font(name=base.name or "Arial", size=base.size or 10,
                         bold=base.bold, color=(BLUE if is_blue else None))

    cl = COLLETTER
    for i, fd in enumerate(funds):
        r = first + i
        setcell(r, "Fund", fd["name"])
        setcell(r, "Strategy", fd.get("strategy", ""))
        setcell(r, "Vintage", str(fd.get("vintage", "") or ""))
        setcell(r, "Commitment", fd["commitment"])
        setcell(r, "Paid-In Capital", fd["paid_in"])
        setcell(r, "Cumulative Distributions", fd.get("distributions", 0))
        setcell(r, "Current NAV", fd["nav"])
        setcell(r, "Net IRR", fd.get("net_irr"))
        setcell(r, "Unfunded Commitment",
                f'=IF(${cl["Commitment"]}{r}="","",'
                f'${cl["Commitment"]}{r}-${cl["Paid-In Capital"]}{r})', formula=True)
        setcell(r, "Total Value",
                f'=IF(${cl["Paid-In Capital"]}{r}="","",'
                f'${cl["Current NAV"]}{r}+${cl["Cumulative Distributions"]}{r})', formula=True)
        setcell(r, "DPI",
                f'=IF(${cl["Paid-In Capital"]}{r}="","",'
                f'IFERROR(${cl["Cumulative Distributions"]}{r}/${cl["Paid-In Capital"]}{r},0))',
                formula=True)
        setcell(r, "TVPI",
                f'=IF(${cl["Paid-In Capital"]}{r}="","",'
                f'IFERROR(${cl["Total Value"]}{r}/${cl["Paid-In Capital"]}{r},0))', formula=True)
    # clear any rows between the funds and the total
    for r in range(last + 1, total_row):
        for c in COL.values():
            ws.cell(row=r, column=c, value=None)
    # total row
    for name in ["Commitment", "Paid-In Capital", "Unfunded Commitment",
                 "Cumulative Distributions", "Current NAV", "Total Value"]:
        ws.cell(row=total_row, column=COL[name],
                value=f"=SUM({cl[name]}{first}:{cl[name]}{last})").number_format = fmt[name]
    ws.cell(row=total_row, column=COL["DPI"],
            value=f'=IF({cl["Paid-In Capital"]}{total_row}=0,"",'
                  f'{cl["Cumulative Distributions"]}{total_row}/{cl["Paid-In Capital"]}{total_row})'
            ).number_format = MULT
    ws.cell(row=total_row, column=COL["TVPI"],
            value=f'=IF({cl["Paid-In Capital"]}{total_row}=0,"",'
                  f'{cl["Total Value"]}{total_row}/{cl["Paid-In Capital"]}{total_row})'
            ).number_format = MULT
    for name in COL:
        ws.cell(row=total_row, column=COL[name]).font = Font(name="Arial", size=10, bold=True)

    if spec.get("as_of"):
        for row in ws.iter_rows(min_row=1, max_row=hdr - 1):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip().startswith("As of"):
                    cell.value = (f"As of {spec['as_of']}     |     Reporting Currency: "
                                  f"{spec.get('reporting_currency', 'USD')}")
    return first, total_row


def write_history(wb, snapshot, quarter_label):
    """Replace the History tab: keep prior quarters, replace/add this quarter."""
    prior = read_existing_history(wb)
    prior[str(quarter_label)] = snapshot          # replace this quarter's block
    if "History" in wb.sheetnames:
        del wb["History"]
    hs = wb.create_sheet("History", 1)
    hs.sheet_view.showGridLines = False
    widths = [12, 32, 16, 15, 15, 16, 16, 15, 15, 9, 9, 10]
    for i, w in enumerate(widths, 1):
        hs.column_dimensions[get_column_letter(i)].width = w

    hs["A1"].value = "Portfolio History — Quarterly Snapshots"
    hs["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY)
    hs["A2"].value = ("A frozen record. Each quarter the skill runs, it logs every fund's "
                      "figures and a portfolio total. Re-running a quarter replaces its block.")
    hs["A2"].font = Font(name="Arial", size=9, italic=True, color="FF595959")

    hrow = 4
    for i, name in enumerate(HISTORY_COLS, 1):
        c = hs.cell(row=hrow, column=i, value=name)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(vertical="center", wrap_text=True)
    hs.row_dimensions[hrow].height = 26

    def putrow(r, quarter, fund, strat, comm, paid, dist, nav, is_total):
        unf, tv = comm - paid, nav + dist
        dpi = dist / paid if paid else 0
        tvpi = tv / paid if paid else 0
        vals = {"Quarter": quarter, "Fund": fund, "Strategy": strat,
                "Commitment": comm, "Paid-In Capital": paid,
                "Unfunded Commitment": unf, "Cumulative Distributions": dist,
                "Current NAV": nav, "Total Value": tv, "DPI": dpi, "TVPI": tvpi}
        for i, name in enumerate(HISTORY_COLS, 1):
            cell = hs.cell(row=r, column=i)
            if name == "Net IRR":
                continue
            cell.value = vals[name]
            if name in HISTORY_FMT:
                cell.number_format = HISTORY_FMT[name]
            cell.font = Font(name="Arial", size=9, bold=is_total)
            if is_total:
                cell.fill = PatternFill("solid", start_color=SHADE)

    r = hrow + 1
    qtotals = []
    for q in sorted(prior, key=quarter_key):
        block = prior[q]
        tc = tp = td = tn = 0
        for fd in block:
            comm = fd["commitment"]; paid = fd["paid_in"]
            dist = fd.get("distributions", 0); nav = fd["nav"]
            putrow(r, q, fd["name"], fd.get("strategy", ""), comm, paid, dist, nav, False)
            irr = hs.cell(row=r, column=12)
            irr.value = fd.get("net_irr")
            irr.number_format = PCT
            irr.font = Font(name="Arial", size=9)
            tc += comm; tp += paid; td += dist; tn += nav
            r += 1
        putrow(r, q, "TOTAL — PORTFOLIO", "", tc, tp, td, tn, True)
        r += 1
        qtotals.append((q, tn + td, tn))      # quarter, total value, NAV
    hs.freeze_panes = "A5"

    # chart-source helper table (portfolio totals by quarter), columns N-P
    hs.cell(row=hrow - 1, column=14,
            value="Chart source — portfolio totals by quarter").font = Font(
        name="Arial", size=8, italic=True, color="FF595959")
    for col, txt in [(14, "Quarter"), (15, "Total Value"), (16, "Current NAV")]:
        c = hs.cell(row=hrow, column=col, value=txt)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", start_color=NAVY)
        hs.column_dimensions[get_column_letter(col)].width = 15
    for i, (q, tv, nav) in enumerate(qtotals):
        rr = hrow + 1 + i
        hs.cell(row=rr, column=14, value=q).font = Font(name="Arial", size=9)
        for col, val in [(15, tv), (16, nav)]:
            cc = hs.cell(row=rr, column=col, value=val)
            cc.number_format = CUR
            cc.font = Font(name="Arial", size=9)
    vals = [v for _, tv, nav in qtotals for v in (tv, nav)]
    return len(qtotals), (min(vals) if vals else 0), (max(vals) if vals else 0)


def nice_axis(lo, hi):
    """A y-axis floor/ceiling that frames the data without starting at zero,
    so quarter-to-quarter movement is readable."""
    span = hi - lo
    pad = max(span * 0.18, hi * 0.03, 1.0)
    step = 10 ** (math.floor(math.log10(max(hi + pad, 10))) - 1) * 5
    axis_min = max(0, math.floor((lo - pad) / step) * step)
    axis_max = math.ceil((hi + pad) / step) * step
    return axis_min, axis_max


def add_history_chart(dashboard_ws, history_ws, n_quarters, lo=0, hi=0):
    """Place a line chart of portfolio Total Value and NAV over time on the
    dashboard, sourced from the History tab's quarter-totals helper table.
    Rebuilt every run, so it extends itself as quarters accumulate."""
    dashboard_ws._charts = []                 # drop any chart from a prior run
    if n_quarters < 1:
        return
    chart = LineChart()
    chart.title = "Portfolio Value Over Time"
    chart.style = 12
    chart.height = 8.5
    chart.width = 22
    last = 4 + n_quarters
    data = Reference(history_ws, min_col=15, max_col=16, min_row=4, max_row=last)
    cats = Reference(history_ws, min_col=14, min_row=5, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.y_axis.title = "USD"
    chart.y_axis.numFmt = '$#,##0,,"M"'
    chart.x_axis.title = "Quarter"
    if hi > 0:
        chart.y_axis.scaling.min, chart.y_axis.scaling.max = nice_axis(lo, hi)
    for s, col in zip(chart.series, ("1F3A5F", "C8862C")):
        s.smooth = False
        s.graphicalProperties = GraphicalProperties(
            ln=LineProperties(solidFill=col, w=32000))
        s.marker = Marker(symbol="circle", size=8,
                          spPr=GraphicalProperties(solidFill=col))
    # anchored below the table and the "How to use" notes — on the front page
    anchor_row = 1
    for row in dashboard_ws.iter_rows(min_col=1, max_col=1):
        if isinstance(row[0].value, str) and row[0].value.strip().startswith("•"):
            anchor_row = max(anchor_row, row[0].row)
    dashboard_ws.add_chart(chart, f"A{anchor_row + 2}")
    # extend the print area so the chart prints on the front page
    last_col = "L"
    if dashboard_ws.print_area:
        m = re.search(r'\$?([A-Z]+)\$?\d+\s*$', str(dashboard_ws.print_area))
        if m:
            last_col = m.group(1)
    dashboard_ws.print_area = f"A1:{last_col}{anchor_row + 24}"


def build_reconciliation(wb, sheet_name, funds, first_data_row, total_row):
    """(Re)build the Reconciliation Key. Covers every fund on the dashboard;
    funds with no fresh statement this run are marked carried-forward."""
    if "Reconciliation Key" in wb.sheetnames:
        del wb["Reconciliation Key"]
    rk = wb.create_sheet("Reconciliation Key")
    rk.sheet_view.showGridLines = False
    rk.column_dimensions['A'].width = 25
    rk.column_dimensions['B'].width = 17
    rk.column_dimensions['C'].width = 96
    q = f"'{sheet_name}'!"
    thin = Side(style="thin", color="FFBFBFBF")

    rk['A1'].value = "Reconciliation Key — Source of Every Figure"
    rk['A1'].font = Font(name="Arial", size=14, bold=True, color=NAVY)
    rk['A2'].value = ("For each fund: where every current roll-up figure came from, the "
                      "unit conversion applied, and any judgment call.")
    rk['A2'].font = Font(name="Arial", size=9, italic=True, color="FF595959")
    rk['A4'].value = "Conventions applied across all funds"
    rk['A4'].font = Font(name="Arial", size=10, bold=True, color=NAVY)
    conventions = [
        "Unfunded Commitment = Commitment - Paid-In Capital (formula). Recallable distributions excluded.",
        "Statements reported in thousands or millions are converted to full dollars before entry.",
        "Where a statement shows a Fund column and an Investor's Share column, the Investor's Share is used.",
        "Where a statement has multiple period columns, Since-Inception figures are used for cumulative amounts.",
        "DPI and TVPI are recomputed by formula; Net IRR is entered as reported and not aggregated.",
    ]
    r = 5
    for t in conventions:
        rk.cell(row=r, column=1, value="-  " + t).font = Font(name="Arial", size=9, color="FF595959")
        r += 1
    r += 1

    carried_notes = {f: "Carried forward — figure unchanged from the last statement processed."
                     for f, _, _ in RECON_FIELDS}
    for i, fd in enumerate(funds):
        rollup_row = first_data_row + i
        fresh = fd.get("_fresh", True)
        rk.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        for col in (1, 2, 3):
            rk.cell(row=r, column=col).fill = PatternFill("solid", start_color=NAVY)
        bc = rk.cell(row=r, column=1, value=fd["name"])
        bc.font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
        bc.alignment = Alignment(vertical="center")
        rk.row_dimensions[r].height = 22
        r += 1
        src = fd.get("source_file", "") if fresh else "— (no new statement this run)"
        units = fd.get("reported_in", "") if fresh else "Carried forward from the prior period."
        for lab, val in [("Source file", src), ("Reported in", units)]:
            rk.cell(row=r, column=1, value=lab).font = Font(name="Arial", size=9, bold=True)
            rk.cell(row=r, column=3, value=val).font = Font(name="Arial", size=9, color="FF595959")
            r += 1
        for col, txt in [(1, "Roll-up field"), (2, "Value"),
                         (3, "Source on the statement / judgment call")]:
            hc = rk.cell(row=r, column=col, value=txt)
            hc.font = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")
            hc.fill = PatternFill("solid", start_color=SUBHEAD)
        rk.row_dimensions[r].height = 18
        r += 1
        notes = fd.get("field_notes", {}) if fresh else carried_notes
        for field, colname, fmt in RECON_FIELDS:
            rk.cell(row=r, column=1, value=field).font = Font(name="Arial", size=9)
            ref = rk.cell(row=r, column=2, value=f"={q}{COLLETTER[colname]}{rollup_row}")
            ref.font = Font(name="Arial", size=9, color=GREEN)
            ref.number_format = fmt
            nc = rk.cell(row=r, column=3, value=notes.get(field, ""))
            nc.font = Font(name="Arial", size=9, color="FF333333")
            nc.alignment = Alignment(wrap_text=True, vertical="top")
            for col in (1, 2, 3):
                rk.cell(row=r, column=col).border = Border(bottom=thin)
            rk.row_dimensions[r].height = 30
            r += 1
        wl = rk.cell(row=r, column=1, value="Watch-outs")
        wl.font = Font(name="Arial", size=9, bold=True, color="FFB35900")
        wl.alignment = Alignment(vertical="top")
        watch = fd.get("watch_outs", "") if fresh else \
            "No new statement this run; figures carried forward from the prior period."
        wc = rk.cell(row=r, column=3, value=watch)
        wc.font = Font(name="Arial", size=9, italic=True, color="FF7A4A00")
        wc.alignment = Alignment(wrap_text=True, vertical="top")
        rk.row_dimensions[r].height = 78
        r += 2

    rk.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    for col in (1, 2, 3):
        rk.cell(row=r, column=col).fill = PatternFill("solid", start_color=NAVY)
    tc = rk.cell(row=r, column=1, value=f"Portfolio Totals  ({len(funds)} funds)")
    tc.font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
    rk.row_dimensions[r].height = 22
    r += 1
    for field, colname, fmt in RECON_FIELDS[:-1]:
        rk.cell(row=r, column=1, value=field).font = Font(name="Arial", size=9, bold=True)
        ref = rk.cell(row=r, column=2, value=f"={q}{COLLETTER[colname]}{total_row}")
        ref.font = Font(name="Arial", size=9, bold=True, color=GREEN)
        ref.number_format = fmt
        for col in (1, 2, 3):
            rk.cell(row=r, column=col).border = Border(bottom=thin)
        r += 1
    rk.cell(row=r, column=1, value="Net IRR").font = Font(name="Arial", size=9, bold=True)
    rk.cell(row=r, column=3, value="Not additive across funds — left blank.").font = Font(
        name="Arial", size=9, italic=True, color="FF333333")


def merge_funds(existing, spec_funds):
    """Existing dashboard funds, updated where a fresh statement matches by name;
    genuinely new funds appended. Funds with no fresh statement carry forward."""
    spec_by_name = {f["name"].strip(): f for f in spec_funds}
    merged, seen = [], set()
    for ef in existing:
        if ef["name"] in spec_by_name:
            m = dict(spec_by_name[ef["name"]]); m["_fresh"] = True
            merged.append(m); seen.add(ef["name"])
        else:
            merged.append(ef)        # carried forward (_fresh already False)
    for sf in spec_funds:
        if sf["name"].strip() not in seen:
            m = dict(sf); m["_fresh"] = True
            merged.append(m)
    return merged


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python build_rollup.py <spec.json>")
    spec = json.load(open(sys.argv[1]))
    spec_funds = spec["funds"]
    if not spec_funds:
        raise SystemExit("ERROR: spec contains no funds.")

    existing_path = spec.get("existing_rollup_path")
    updating = bool(existing_path) and os.path.exists(existing_path)
    base = existing_path if updating else spec["template_path"]
    wb = openpyxl.load_workbook(base)
    ws, hdr, COL = find_rollup_sheet(wb)
    for name, c in COL.items():
        COLLETTER[name] = get_column_letter(c)
    total_row = find_total_row(ws, hdr, COL["Fund"])

    existing = read_existing_funds(ws, hdr, COL, total_row) if updating else []
    merged = merge_funds(existing, spec_funds)

    first, total_row = write_front_page(ws, hdr, COL, total_row, merged, spec)
    n_quarters, lo, hi = write_history(
        wb, merged, spec.get("quarter_label", spec.get("as_of", "")))
    add_history_chart(ws, wb["History"], n_quarters, lo, hi)
    build_reconciliation(wb, ws.title, merged, first, total_row)

    out = spec["output_path"]
    tmp = out + ".tmp"
    wb.save(tmp)
    os.replace(tmp, out)
    recalc(out)

    # ---- verification (independent of the spreadsheet's own formulas) ----
    mode = "UPDATED existing roll-up" if updating else "CREATED new roll-up"
    print(f"\n{mode}: {out}")
    print(f"Tabs: Portfolio Roll-Up (dashboard) · History · Reconciliation Key")
    fresh = [f["name"] for f in merged if f.get("_fresh")]
    carried = [f["name"] for f in merged if not f.get("_fresh")]
    print(f"Funds refreshed this run: {', '.join(fresh) if fresh else '(none)'}")
    if carried:
        print(f"Funds carried forward:   {', '.join(carried)}")
    print(f"\n{'Fund':34}{'Commit':>14}{'Paid-In':>14}{'Unfund':>13}"
          f"{'Distrib':>13}{'NAV':>14}{'TotVal':>14}{'DPI':>8}{'TVPI':>8}")
    tc = tp = tu = td = tn = 0
    problems = []
    for fd in merged:
        comm, paid = fd["commitment"], fd["paid_in"]
        dist, nav = fd.get("distributions", 0), fd["nav"]
        unf, tv = comm - paid, nav + dist
        dpi = dist / paid if paid else 0
        tvpi = tv / paid if paid else 0
        tc += comm; tp += paid; tu += unf; td += dist; tn += nav
        print(f"{fd['name'][:34]:34}{comm:>14,.0f}{paid:>14,.0f}{unf:>13,.0f}"
              f"{dist:>13,.0f}{nav:>14,.0f}{tv:>14,.0f}{dpi:>7.2f}x{tvpi:>7.2f}x")
        if fd.get("_fresh"):
            for label, computed, reported in [("DPI", dpi, fd.get("reported_dpi")),
                                              ("TVPI", tvpi, fd.get("reported_tvpi"))]:
                if reported is not None and abs(computed - reported) > 0.02:
                    problems.append(f"{fd['name']}: recomputed {label} {computed:.2f}x differs "
                                    f"from statement's reported {reported:.2f}x — re-check extraction.")
    print(f"{'TOTAL — PORTFOLIO':34}{tc:>14,.0f}{tp:>14,.0f}{tu:>13,.0f}"
          f"{td:>13,.0f}{tn:>14,.0f}{tn+td:>14,.0f}"
          f"{td/tp if tp else 0:>7.2f}x{(tn+td)/tp if tp else 0:>7.2f}x")
    if problems:
        print("\nWARNINGS — review before sending:")
        for p in problems:
            print("  !", p)
    else:
        print("\nAll DPI / TVPI cross-checks tie out to the statements.")


if __name__ == "__main__":
    main()
